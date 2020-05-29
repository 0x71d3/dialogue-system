import glob
import os
import sys
from openpyxl import load_workbook

xlsx_paths = glob.glob(sys.argv[1] + '/**/*.xlsx', recursive=True)

for xlsx_path in xlsx_paths[1:]:
    wb = load_workbook(xlsx_path)
    ws = wb.active

    data = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        cell_range = row[3:8]
        if None in cell_range:
            break
        data.append(tuple(map(str, cell_range)))

    wb.close()

    xlsx_name = os.path.basename(xlsx_path)
    tsv_name = os.path.splitext(xlsx_name)[0] + '.tsv'
    tsv_path = os.path.join(sys.argv[2], tsv_name)

    with open(tsv_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join('\t'.join(line) for line in data))
